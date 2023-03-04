from django.conf import settings
from .models import Listing
import datetime
from datetime import date
from datetime import datetime
from django.core.paginator import Paginator
from .my_functions import update_next_check, update_last_checked, update_date_appearance, update_next_check_for_one
from .choices import units_choices, blocks_choices, control_users, types_choices, special_types_choices, \
     measure_units_choices, realtors_choices, is_active_choices, intervals_choices
from django.shortcuts import render, redirect, get_object_or_404
# from .forms import ListingForm, PostFileForm
# from .forms import ListingForm, FileForm
from .forms import ListingForm, ReportForm, FullReportForm
from .models import Report, FullReport

from django.contrib import messages

from openpyxl import Workbook, load_workbook

import pathlib

from .filters import ListingFilter
import csv
from django.shortcuts import render

# from django_tables2 import SingleTableView
# from .tables import ListingTable
# from django_tables2.views import SingleTableMixin


# Create your views here.
# class ListingView(LoginRequiredMixin, View):
#     def get(self, request):
#         listings = Listing.objects.all()
#         context = {'listings': listings}
#         return render(request, 'listings/listings.html', context)


# def table_test(request): # commented 25/02/2023 ################################
#     context = {}
#     return render(request, 'listings/table-test.html', context)
# class ListingListView(ListView):

# WITH CLASS
# class ListingListView(SingleTableView):
#     model = Listing
#     table_class = ListingTable
#     template_name = 'listings/table-test.html'

# WITH FUNCTION
# def table_test(request):
#     listings = Listing.objects.all().order_by('tag')
#     update_next_check(listings)  # δεν λειτουργει εδω
#     # table = ListingTable(update_next_check(Listing.objects.all().order_by('tag')))
#     table = ListingTable(listings)
#     # table = ListingTable(update_next_check(listings))
#     # table = ListingTable(update_next_check(listings))
#
#     listings = update_next_check(Listing.objects.all().order_by('tag'))
#     # myFilter = ListingFilter(request.GET, queryset=listings)
#     myFilter = ListingFilter(request.GET, queryset=listings)
#     listings = myFilter.qs
#
#     table.paginate(page=request.GET.get("page", 1), per_page=25)
#     update_next_check(myFilter.qs)
#     update_next_check(listings)
#
#     context = {
#         'table': table,
#         'myFilter': myFilter,
#         # 'listings': listings,
#     }
#     return render(request, 'listings/table-test.html', context)


# index with django_filters
def index(request):

    # listings = Listing.objects.order_by('unit')
    try:
        order_by = request.GET.get('order_by', 'tag')
        listings = Listing.objects.order_by(order_by)
    except ProgrammingError:
        listings = []

    # listings = Listing.objects.order_by('tag')
    # listings = Listing.objects.all().order_by('tag')
    update_next_check(listings)


    ###################
    # myFilter = ListingFilter(request.GET, queryset=listings)
    # listings = myFilter.qs
    update_next_check(listings)
    ##################

    paginator = Paginator(listings, 25)
    page = request.GET.get('page')
    paged_listings = paginator.get_page(page)

    update_next_check(listings)
    update_date_appearance(listings)
    listings = Listing.objects.all()
    for listing in listings:
        update_next_check_for_one(listing)
        listing.save()
        # print(listing.tag, listing.id, listing.next_check)

    context = {
        'listings': paged_listings,
        # 'myFilter': myFilter,
        # 'listings': listings,
    }

    return render(request, 'listings/listings.html', context)


def listing(request, listing_id):
    try:
        listing = get_object_or_404(Listing, pk=listing_id)
    except ProgrammingError:
        listings = []

    fields = Listing._meta.get_fields()
    reports = Report.objects.all().filter(listing_id=listing_id)
    full_reports = FullReport.objects.all().filter(listing_id=listing_id)

    context = {
        'listing': listing,
        'reports': reports,
        'full_reports': full_reports,
    }
    return render(request, 'listings/listing.html', context)

# class SingleListingView(DetailView):
#     template_name = 'listings/listing.html'
#     model = Listing


def search(request):
    try:
        listings = Listing.objects.all()
    except ProgrammingError:
        listings = []
    update_next_check(listings)
    queryset_list = update_next_check(Listing.objects.order_by('tag'))

    # Keywords
    if 'keywords' in request.GET:
        keywords = request.GET['keywords']
        if keywords:
            queryset_list = queryset_list.filter(tag__icontains=keywords)
            update_next_check(queryset_list)

    # Last Checked from
    if 'last_checked_from' in request.GET:
        last_checked_from = request.GET['last_checked_from']
        if last_checked_from:
            queryset_list = queryset_list.filter(last_checked__gte=last_checked_from)
            update_next_check(queryset_list)

    # Is Active
    if 'is_active' in request.GET:
        is_active = request.GET['is_active']
        if is_active:
            queryset_list = queryset_list.filter(is_active__iexact=is_active)
            update_next_check(queryset_list)

    # Last Checked to
    if 'last_checked_to' in request.GET:
        last_checked_to = request.GET['last_checked_to']
        if last_checked_to:
            queryset_list = queryset_list.filter(last_checked__lte=last_checked_to)
            update_next_check(queryset_list)

    # Next Check from
    if 'next_check_from' in request.GET:
        next_check_from = request.GET['next_check_from']
        if next_check_from:
            queryset_list = queryset_list.filter(next_check__gte=next_check_from)
            update_next_check(queryset_list)

    # Next Check to
    if 'next_check_to' in request.GET:
        next_check_to = request.GET['next_check_to']
        if next_check_to:
            queryset_list = queryset_list.filter(next_check__lte=next_check_to)
            update_next_check(queryset_list)

    # Block
    if 'block' in request.GET:
        block = request.GET['block']
        if block:
            queryset_list = queryset_list.filter(block__iexact=block)
            update_next_check(queryset_list)

    # Interval
    if 'interval' in request.GET:
        interval = request.GET['interval']
        if interval:
            queryset_list = queryset_list.filter(interval__iexact=interval)
            update_next_check(queryset_list)

    # Unit
    if 'unit' in request.GET:
        unit = request.GET['unit']
        if unit:
            queryset_list = queryset_list.filter(unit__iexact=unit)
            update_next_check(queryset_list)

    # Type
    if 'type' in request.GET:
        type = request.GET['type']
        if type:
            queryset_list = queryset_list.filter(type__iexact=type)
            update_next_check(queryset_list)

    # special_Type
    if 'special_type' in request.GET:
        special_type = request.GET['special_type']
        if special_type:
            queryset_list = queryset_list.filter(special_type__iexact=special_type)
            update_next_check(queryset_list)

    # Measure_Units
    if 'units' in request.GET:
        units = request.GET['units']
        if units:
            queryset_list = queryset_list.filter(units__iexact=units)
            update_next_check(queryset_list)

    # Realtor
    if 'realtor' in request.GET:
        realtor = request.GET['realtor']
        if realtor:
            queryset_list = queryset_list.filter(realtor__name__iexact=realtor)
            update_next_check(queryset_list)

    update_date_appearance(queryset_list)

    context = {
        'units_choices': units_choices,
        'blocks_choices': blocks_choices,
        'types_choices': types_choices,
        'special_types_choices': special_types_choices,
        'measure_units_choices': measure_units_choices,
        'realtors_choices': realtors_choices,
        'is_active_choices': is_active_choices,
        'intervals_choices': intervals_choices,
        'listings': queryset_list,
        'values': request.GET
    }
    return render(request, 'listings/search.html', context)


#################################################################################
def edit(request, listing_id):
    # context = {}
    # fetch the object related to passed id
    try:
        listing = get_object_or_404(Listing, id=listing_id)
    except ProgrammingError:
        listings = []
    context = {
        'listing': listing,
    }
    form = ListingForm(request.POST or None, instance=listing)
    # save the data from the form and redirect to detail_view
    if form.is_valid():
        update_next_check_for_one(listing)
        form.save()
        messages.success(request, f'{listing} was UPDATED')
        # return redirect("listing")
        return render(request, "listings/listing.html", context)

    # add form dictionary to context
    context["form"] = form

    # return render(request, 'listings/listing.html', context)
    return render(request, "listings/edit.html", context)

############################################################################################################


def insert(request):
    # dictionary for initial data with field names as keys
    context = {}

    # add the dictionary during initialization
    form = ListingForm(request.POST or None)
    if form.is_valid():
        if request.user.is_authenticated:
            form.save()
            messages.success(request, f"{Listing.objects.filter().order_by('id').last()} was CREATED SUCCESSFULLY")
            return redirect("listings")
        else:
            messages.error(request, "Listing was NOT created. LOOK FOR PROBLEMS")
            return redirect("listings")

    context['form'] = form
    return render(request, "listings/insert.html", context)


############################################################################################################

# class ListingDeleteView(DeleteView):
#     # looks for model_confirm_delete.html
#     model = Listing
#     success_url = reverse_lazy('listings')


def remove(request, listing_id):

    # fetch the object related to passed id
    try:
        listing = get_object_or_404(Listing, id=listing_id)
    except ProgrammingError:
        listings = []
    context = {
        'listing': listing,
    }
    if request.method == 'POST':
        if request.user.is_authenticated:
            if request.user.username in control_users:
                print(request.user.username in control_users)
                # delete object
                listing.delete()
                # after deleting redirect to listings page
                messages.success(request, f'{listing} was deleted')
                return redirect("listings")
            else:
                messages.error(request, f'No credentials for deleting {listing}')
                return redirect("listings")

    return render(request, "listings/listing_confirm_delete.html", context)


###################################################################################################
#################################################################################################

    # return render(request, 'listings/create-report.html',)
###################################################################################################
####################################################################################################
def upload_report(request, listing_id):
    # context = {}
    try:
        listing = get_object_or_404(Listing, id=listing_id)
    except ProgrammingError:
        listings = []
    context = {
            'listing': listing,
        }
    if request.method == 'POST':
        form = ReportForm(request.POST, request.FILES)
        if form.is_valid():
            update_last_checked(listing)
            listing.save()
            form.save()
            messages.success(request, f'{listing} was UPDATED')
            # return redirect('listings')
            return redirect('listing', listing_id)
            # return render(request, 'listings/listing.html', context)
    else:
        form = ReportForm()
        context = {
            'listing': listing,
            'form': form,
        }

    return render(request, 'listings/upload-report.html', context)


#######################################################################################################
# λειτουργει χωρις το upload του file
##
#######################################################################################################
# def create_report(request, listing_id):
#     listing = get_object_or_404(Listing, id=listing_id)
#     context = {
#         'listing': listing,
#     }
#     old_history = listing.history
#     form = CreateReportForm(request.POST or None, instance=listing)
#     if request.method == 'POST':
#         if form.is_valid():
#             new_history = request.POST.get('history')
#             listing.history = old_history + '\n\n' + new_history
#             wb = Workbook()
#             ws = wb.active
#             ws['A4'] = listing.tag
#             ws['A6'] = listing.history
#             wb.save('my_first_report6.xlsx')
#             #file = 'my_first_report6.xlsx'
#             update_last_checked(listing)
#             listing.save()
#             # form.save()
#             # messages.success(request, f'{listing} report was created')
#             messages.success(request, f'{listing} report was created')
#             return redirect('listings')
#             # return render(request, 'listings/listing.html', context)
#     else:
#         form = CreateReportForm()
#         context = {
#             'listing': listing,
#             'form': form,
#         }
#
#     return render(request, 'listings/create-report.html', context)

##########################################################################################################
#    TEST   ΛΕΙΤΟΥΡΓΕΙ ΟΧΙ ΤΕΛΕΙΑ ΟΧΙ PYTHONIC
##########################################################################################################
# def create_report(request, listing_id):
#     listing = get_object_or_404(Listing, id=listing_id)
#     context = {
#         'listing': listing,
#     }
#     old_history = listing.history
#     form = CreateReportForm(request.POST or None, instance=listing)
#     if request.method == 'POST':
#         if form.is_valid():
#             new_history = request.POST.get('history')
#             listing.history = old_history + '\n\n' + new_history
#             wb = Workbook()
#             ws = wb.active
#             ws['A4'] = listing.tag
#             ws['A6'] = listing.history
#             wb.save('my_first_report16.xlsx')
#             file = 'my_first_report16.xlsx'
#             update_last_checked(listing)
#             listing.save()
#             new_report = Report(listing=listing, file=file)
#             new_report.save()
#             context = {
#                 'listing': listing,
#                 # 'file': file,
#                 # 'new_report': new_report,
#             }
#             messages.success(request, f'{listing} report was created and uploaded!!!')
#             # return redirect('listings')
#             return redirect('listing', listing_id)
#             # return render(request, 'listings/listings.html', context)
#
#     else:
#         form = CreateReportForm()
#         context = {
#             'listing': listing,
#             'form': form,
#         }
#     return render(request, 'listings/create-report.html', context)

########################################################################################################
# TEST 2
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
########################################################################################################
def create_report(request, listing_id):
    try:
        listing = get_object_or_404(Listing, pk=listing_id)
    except ProgrammingError:
        listings = []
    context = {
        'listing': listing,
    }
    old_history = listing.history

    if request.method == 'POST':
        # form = FullReportForm(request.POST or None, instance=listing)
        form = FullReportForm(request.POST)
        if form.is_valid():
            new_report = form.save()
            comments = new_report.comments
            listing.history = old_history + '\n\n' + comments
            wb = Workbook()
            # wb = load_workbook(filename=r'C:\Users\ALEXIS\OneDrive\PYTHON-LESSONS\DJANGO-ALL\instruments_project\media\excel_files\Certificate_sample.xlsx')
            filepath = settings.MEDIA_ROOT + '/excel_files/Certificate_sample.xlsx'
            wb = load_workbook(filename=filepath)
            ws = wb.active
            ws['F7'] = listing.tag
            ws['F9'] = listing.description
            ws['Z7'] = listing.manufacturer
            ws['Z8'] = listing.model
            ws['Z9'] = listing.serial_no
            ws['Z11'] = listing.lrv
            ws['AB11'] = listing.urv
            ws['AC12'] = listing.units
            ws['J15'] = listing.detailed_type
            ws['J17'] = listing.interval
            ws['G24'] = new_report.output_0_before_cal
            ws['G25'] = new_report.output_25_before_cal
            ws['G26'] = new_report.output_50_before_cal
            ws['G27'] = new_report.output_75_before_cal
            ws['G28'] = new_report.output_100_before_cal
            ws['X24'] = new_report.output_0_after_cal
            ws['X25'] = new_report.output_25_after_cal
            ws['X26'] = new_report.output_50_after_cal
            ws['X27'] = new_report.output_75_after_cal
            ws['X28'] = new_report.output_100_after_cal
            ws['G41'] = new_report.manifold_correct
            ws['G42'] = new_report.ferrules_correct
            ws['G43'] = new_report.mounting_correct
            ws['T41'] = new_report.cable_gland_correct
            ws['T46'] = new_report.insulation_correct
            ws['AC45'] = new_report.general_state_correct
            ws['F55'] = str(new_report.date_created)
            ws['O55'] = str(listing.realtor)
            ws['B48'] = comments
            # filepath = 'Certificate_sample10009.xlsx'
            # now = str(datetime.datetime.now())
            today = date.today()
            d1 = str(today.strftime("%Y/%m/%d"))
            # print(today)
            # filepath1 = 'C:/Users/ALEXIS/OneDrive/PYTHON-LESSONS/DJANGO-ALL/instruments_project/media/excel_files'
            # filepath = filepath1 + "\\" + listing.tag + "\\" + d1 + '.xlsx' # 'Certificate_sample14A.xlsx'
            # filepath = settings.MEDIA_ROOT + "/" + listing.tag + '.xlsx' #
            pathlib.Path(settings.MEDIA_ROOT + '/report_files/' + d1 + '/').mkdir(parents=True, exist_ok=True)
            filepath = settings.MEDIA_ROOT + '/report_files/' + d1 + '/' + listing.tag + '.xlsx'
            wb.save(filename=filepath)
            wb.close()
            # new_report.file.save(new_name, File(f))
            # new_report.file = Path(filepath)
            # f = open('Certificate_sample11A.xlsx', 'rb').read()

            # new_report.file.save(new_name, File(f))
            # new_report = FullReport.objects.create(listing=listing)
            # new_report = FullReport(listing=listing)
            new_report.listing = listing
            # new_report.file = 'reports_files/2023/01/31/Certificate_sample10A.xlsx'
            # new_report.file = filepath
            # new_report.file = listing.tag + '.xlsx'
            new_report.file = 'report_files/' + d1 + '/' + listing.tag + '.xlsx'
            new_report.save()

            # new_report.update(listing=listing)
            update_last_checked(listing)
            listing.save()

            my_report = FullReport.objects.all().filter(listing_id=listing_id).order_by('-id')[0]
            # my_report = FullReport.objects.get(listing_id)
            # my_report = get_object_or_404(FullReport, pk=listing.fullreport.id)
            # new_report.file.path = 'Certificate_sample14A.xlsx'
            # new_report.save()
            # path = Path('Certificate_sample12A.xlsx')
            # with path.open(mode='rb') as f:
            #     new_report.file = File(f, name=path.name)
            # new_report.save()
            ###############################################################################################
            # alli mia prospaueia
            # f = open(r'Certificate_sample13A.xlsx')
            # myfile = File(f)
            #
            # new_report.file.save(r'Certificate_sample133333333A.xlsx', myfile)

            # from django.core.files import File
            # from openpyxl import Workbook
            # from openpyxl.writer.excel import save_virtual_workbook
            #
            # class Order1(View):
            #     def post(self, request, *args, **kwargs):
            #         form = self.form_class(request.POST)
            #         if form.is_valid():
            #             obj = form.save(commit=False)
            #
            #             # Creating sheet
            #             book = Workbook()
            #             sheet = book.active
            #             sheet['A1'] = 56
            #             sheet['A2'] = 43
            #             now = time.strftime("%x")
            #             sheet['A3'] = now
            #
            #             # Calling model field instance - What im doing wrong here?
            #             book.save("sample.xlsx")
            #             with open('sample.xlsx', 'rb') as f:
            #                 obj.attachment.save("sample.xlsx", File(f), save=False)
            #
            #             # Saving model instance
            #             obj.save()
            #
            #             # Some return - required for AJAX
            #             return JsonResponse({"status": "OK"})

            #################################################################################################

            # new_report.file = f
            # new_report.listing = listing

            # print(new_report.file)

            # with NamedTemporaryFile() as tmp:
            #     wb.save(tmp.name)
            #     tmp.seek(0)
            #     stream = tmp.read()
            # print(type(stream))
            # print(stream)
            # now = str(datetime.datetime.now())
            # today = date.today()
            # d1 = str(today.strftime("%Y/%m/%d"))
            # print(today)
            # wb.save(r'D:\folder\folder\folder\Filename.xlsx')
            # wb.save(filename=filepath)
            # file = 'my_certificate100.xlsx'
            # new_report = FullReport(listing=listing)#, file=wb)
            # new_report.file = wb
            # wb.close()

            # context = {
            #     'listing': listing,
            #     #'form': form,
            #     'new_report': new_report,
            # }
            messages.success(request, f'{listing}: FULL report {new_report} was created and uploaded!!!')
            return redirect('listing', listing_id)
            # return render(request, 'listings/listing.html', context)

    else:
        form = FullReportForm()
        context = {
            'listing': listing,
            'form': form,
        }
    return render(request, 'listings/create-report.html', context)

################################################################################################


# https://stackoverflow.com/questions/57183002/filefield-not-working-with-arrayfield-in-django
# https://code.djangoproject.com/attachment/ticket/25756/multiple.py

################################################################################################################



